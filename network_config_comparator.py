"""
Network Configuration Comparator
Compares pre and post update network command outputs and generates comparison reports
"""

import json
from typing import Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from command_segmenter import (
    read_mx80_show_commands, 
    extract_show_arp_output,
    extract_show_vrrp_summary_output
)
from mx80_parser_engine import parse_show_arp_no_resolve, parse_show_vrrp_summary


# ============================================================================
# PHASE 1: Parse both files and extract JSONs
# ============================================================================

def parse_file(file_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Parse a single file and return all command outputs as JSONs
    
    Args:
        file_path: Path to the text file containing command outputs
        
    Returns:
        Dictionary with command names as keys and their JSON outputs as values
    """
    raw_output = read_mx80_show_commands(file_path)
    
    results = {}
    
    # Parse show arp no-resolve
    arp_output = extract_show_arp_output(raw_output)
    arp_result = parse_show_arp_no_resolve(arp_output)
    results['show_arp_no_resolve'] = arp_result.to_dict()
    
    # Parse show vrrp summary
    vrrp_output = extract_show_vrrp_summary_output(raw_output)
    vrrp_result = parse_show_vrrp_summary(vrrp_output)
    results['show_vrrp_summary'] = vrrp_result.to_dict()
    
    return results


def parse_both_files(pre_file: str, post_file: str) -> Tuple[Dict, Dict]:
    """
    Parse both pre and post update files
    
    Args:
        pre_file: Path to pre-update file
        post_file: Path to post-update file
        
    Returns:
        Tuple of (pre_update_data, post_update_data)
    """
    print("Parsing pre-update file...")
    pre_data = parse_file(pre_file)
    
    print("Parsing post-update file...")
    post_data = parse_file(post_file)
    
    return pre_data, post_data


# ============================================================================
# PHASE 2: Deep comparison logic
# ============================================================================

def compare_values(pre_val: Any, post_val: Any) -> str:
    """
    Compare two values and return color code
    
    Args:
        pre_val: Value from pre-update
        post_val: Value from post-update
        
    Returns:
        "green" if exact match, "red" if different
    """
    if pre_val == post_val:
        return "green"
    return "red"


def compare_arp_entries(pre_entries: List[Dict], post_entries: List[Dict]) -> List[Dict]:
    """
    Compare ARP entries entry-by-entry based on IP address as key
    
    Args:
        pre_entries: List of ARP entries from pre-update
        post_entries: List of ARP entries from post-update
        
    Returns:
        List of comparison results with color codes
    """
    # Create dictionaries indexed by IP address for easy lookup
    pre_dict = {entry['ip_address']: entry for entry in pre_entries}
    post_dict = {entry['ip_address']: entry for entry in post_entries}
    
    # Get all unique IP addresses from both
    all_ips = set(pre_dict.keys()) | set(post_dict.keys())
    
    comparison_entries = []
    
    for ip in sorted(all_ips):
        pre_entry = pre_dict.get(ip)
        post_entry = post_dict.get(ip)
        
        if pre_entry and post_entry:
            # Entry exists in both - compare field by field
            comp_entry = {
                "ip_address": compare_values(pre_entry['ip_address'], post_entry['ip_address']),
                "mac_address": compare_values(pre_entry['mac_address'], post_entry['mac_address']),
                "interface": compare_values(pre_entry['interface'], post_entry['interface']),
                "flags": compare_values(pre_entry['flags'], post_entry['flags'])
            }
        elif pre_entry and not post_entry:
            # Entry deleted (only in pre)
            comp_entry = {
                "ip_address": "red",
                "mac_address": "red",
                "interface": "red",
                "flags": "red",
                "status": "deleted"
            }
        else:
            # Entry added (only in post)
            comp_entry = {
                "ip_address": "red",
                "mac_address": "red",
                "interface": "red",
                "flags": "red",
                "status": "added"
            }
        
        comparison_entries.append(comp_entry)
    
    return comparison_entries


def compare_vrrp_entries(pre_entries: List[Dict], post_entries: List[Dict]) -> List[Dict]:
    """
    Compare VRRP entries entry-by-entry based on interface+group as key
    
    Args:
        pre_entries: List of VRRP entries from pre-update
        post_entries: List of VRRP entries from post-update
        
    Returns:
        List of comparison results with color codes
    """
    # Create dictionaries indexed by interface+group for easy lookup
    pre_dict = {(entry['interface'], entry['group']): entry for entry in pre_entries}
    post_dict = {(entry['interface'], entry['group']): entry for entry in post_entries}
    
    # Get all unique interface+group combinations
    all_keys = set(pre_dict.keys()) | set(post_dict.keys())
    
    comparison_entries = []
    
    for key in sorted(all_keys):
        pre_entry = pre_dict.get(key)
        post_entry = post_dict.get(key)
        
        if pre_entry and post_entry:
            # Entry exists in both - compare field by field
            comp_entry = {
                "interface": compare_values(pre_entry['interface'], post_entry['interface']),
                "state": compare_values(pre_entry['state'], post_entry['state']),
                "group": compare_values(pre_entry['group'], post_entry['group']),
                "vr_state": compare_values(pre_entry['vr_state'], post_entry['vr_state']),
                "vr_mode": compare_values(pre_entry['vr_mode'], post_entry['vr_mode']),
                "addresses": compare_values(pre_entry['addresses'], post_entry['addresses'])
            }
        elif pre_entry and not post_entry:
            # Entry deleted
            comp_entry = {
                "interface": "red",
                "state": "red",
                "group": "red",
                "vr_state": "red",
                "vr_mode": "red",
                "addresses": "red",
                "status": "deleted"
            }
        else:
            # Entry added
            comp_entry = {
                "interface": "red",
                "state": "red",
                "group": "red",
                "vr_state": "red",
                "vr_mode": "red",
                "addresses": "red",
                "status": "added"
            }
        
        comparison_entries.append(comp_entry)
    
    return comparison_entries


def compare_show_arp(pre_arp: Dict, post_arp: Dict) -> Dict:
    """
    Compare show arp no-resolve outputs
    
    Args:
        pre_arp: Pre-update ARP data
        post_arp: Post-update ARP data
        
    Returns:
        Comparison dictionary with color codes
    """
    comparison = {
        "total_entries": compare_values(pre_arp['total_entries'], post_arp['total_entries']),
        "entries": compare_arp_entries(pre_arp['entries'], post_arp['entries'])
    }
    
    return comparison


def compare_show_vrrp(pre_vrrp: Dict, post_vrrp: Dict) -> Dict:
    """
    Compare show vrrp summary outputs
    
    Args:
        pre_vrrp: Pre-update VRRP data
        post_vrrp: Post-update VRRP data
        
    Returns:
        Comparison dictionary with color codes
    """
    comparison = {
        "entries": compare_vrrp_entries(pre_vrrp['entries'], post_vrrp['entries'])
    }
    
    return comparison


def generate_comparison(pre_data: Dict, post_data: Dict) -> Dict:
    """
    Generate complete comparison for all commands
    
    Args:
        pre_data: Pre-update data for all commands
        post_data: Post-update data for all commands
        
    Returns:
        Complete comparison dictionary with color codes
    """
    comparison = {}
    
    # Compare show arp no-resolve
    if 'show_arp_no_resolve' in pre_data and 'show_arp_no_resolve' in post_data:
        comparison['show_arp_no_resolve'] = compare_show_arp(
            pre_data['show_arp_no_resolve'],
            post_data['show_arp_no_resolve']
        )
    
    # Compare show vrrp summary
    if 'show_vrrp_summary' in pre_data and 'show_vrrp_summary' in post_data:
        comparison['show_vrrp_summary'] = compare_show_vrrp(
            pre_data['show_vrrp_summary'],
            post_data['show_vrrp_summary']
        )
    
    return comparison


# ============================================================================
# PHASE 3: Excel Export with Color Coding
# ============================================================================

def determine_overall_color(comparison_data: Dict) -> str:
    """
    Determine overall color for entire command based on comparison data
    Recursively check if ANY field is red
    
    Args:
        comparison_data: Comparison dictionary with color codes
        
    Returns:
        "green" if all fields match, "red" if any field differs
    """
    if isinstance(comparison_data, dict):
        for key, value in comparison_data.items():
            if isinstance(value, str) and value == "red":
                return "red"
            elif isinstance(value, (dict, list)):
                if determine_overall_color(value) == "red":
                    return "red"
    elif isinstance(comparison_data, list):
        for item in comparison_data:
            if isinstance(item, str) and item == "red":
                return "red"
            elif isinstance(item, (dict, list)):
                if determine_overall_color(item) == "red":
                    return "red"
    elif isinstance(comparison_data, str):
        return comparison_data
    
    return "green"


def get_cell_color(color_code: str):
    """
    Get the fill color based on color code
    
    Args:
        color_code: "green" or "red"
        
    Returns:
        PatternFill object
    """
    if color_code == "green":
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    else:  # red
        return PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red


def write_json_to_cells(ws, start_row, start_col, data, indent=0):
    """
    Recursively write JSON data to Excel cells
    
    Args:
        ws: Worksheet object
        start_row: Starting row number
        start_col: Starting column number
        data: JSON data to write
        indent: Indentation level for nested data
        
    Returns:
        Next available row number
    """
    current_row = start_row
    
    if isinstance(data, dict):
        for key, value in data.items():
            # Write the key
            key_cell = ws.cell(row=current_row, column=start_col)
            key_cell.value = "  " * indent + str(key) + ":"
            key_cell.font = Font(bold=True)
            key_cell.alignment = Alignment(horizontal='left', vertical='top')
            
            # Handle the value
            if isinstance(value, (dict, list)):
                current_row += 1
                current_row = write_json_to_cells(ws, current_row, start_col, value, indent + 1)
            else:
                # Write value in the same row
                value_cell = ws.cell(row=current_row, column=start_col)
                value_cell.value = "  " * indent + str(key) + ": " + str(value)
                value_cell.alignment = Alignment(horizontal='left', vertical='top')
                current_row += 1
                
    elif isinstance(data, list):
        for idx, item in enumerate(data):
            if isinstance(item, (dict, list)):
                # Write list item header
                header_cell = ws.cell(row=current_row, column=start_col)
                header_cell.value = "  " * indent + f"[{idx}]"
                header_cell.font = Font(italic=True)
                header_cell.alignment = Alignment(horizontal='left', vertical='top')
                current_row += 1
                current_row = write_json_to_cells(ws, current_row, start_col, item, indent + 1)
            else:
                item_cell = ws.cell(row=current_row, column=start_col)
                item_cell.value = "  " * indent + str(item)
                item_cell.alignment = Alignment(horizontal='left', vertical='top')
                current_row += 1
    else:
        # Scalar value
        cell = ws.cell(row=current_row, column=start_col)
        cell.value = "  " * indent + str(data)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        current_row += 1
    
    return current_row


def export_to_excel(pre_data: Dict, post_data: Dict, comparison: Dict, output_file: str = "network_comparison.xlsx"):
    """
    Export comparison results to Excel file
    
    Args:
        pre_data: Pre-update data
        post_data: Post-update data
        comparison: Comparison data with color codes
        output_file: Output Excel file name
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Network Comparison"
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    
    # Headers
    ws['A1'] = "Pre-Update"
    ws['A1'].font = Font(bold=True, size=14)
    ws['B1'] = "Post-Update"
    ws['B1'].font = Font(bold=True, size=14)
    ws['C1'] = "Status"
    ws['C1'].font = Font(bold=True, size=14)
    
    current_row = 3
    
    # Process each command
    for command_name in pre_data.keys():
        # Determine overall color for this command
        overall_color = determine_overall_color(comparison[command_name])
        
        # Command header
        ws.cell(row=current_row, column=1).value = f"=== {command_name} ==="
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="0000FF")
        ws.cell(row=current_row, column=2).value = f"=== {command_name} ==="
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=12, color="0000FF")
        ws.cell(row=current_row, column=3).value = f"=== {command_name} ==="
        ws.cell(row=current_row, column=3).font = Font(bold=True, size=12, color="0000FF")
        current_row += 1
        
        start_row = current_row
        
        # Write pre-update data in column 1
        end_row_pre = write_json_to_cells(ws, start_row, 1, pre_data[command_name])
        
        # Write post-update data in column 2
        end_row_post = write_json_to_cells(ws, start_row, 2, post_data[command_name])
        
        # Write single colored cell in column 3 for comparison status
        status_cell = ws.cell(row=start_row, column=3)
        status_text = "MATCH" if overall_color == "green" else "CHANGED"
        status_cell.value = status_text
        status_cell.fill = get_cell_color(overall_color)
        status_cell.font = Font(bold=True, size=12)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Move to next command section
        current_row = max(end_row_pre, end_row_post) + 2
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n✓ Excel file saved: {output_file}")


# ============================================================================
# TEST ALL PHASES
# ============================================================================

if __name__ == "__main__":
    # Test Phase 1
    print("=" * 70)
    print("PHASE 1: Parsing both files")
    print("=" * 70)
    
    pre_data, post_data = parse_both_files("pre_update.txt", "post_update.txt")
    
    print("\n--- PRE-UPDATE DATA ---")
    print(json.dumps(pre_data, indent=2))
    
    print("\n--- POST-UPDATE DATA ---")
    print(json.dumps(post_data, indent=2))
    
    print("\n✓ Phase 1 Complete: Both files parsed successfully")
    
    # Test Phase 2
    print("\n" + "=" * 70)
    print("PHASE 2: Generating comparison")
    print("=" * 70)
    
    comparison = generate_comparison(pre_data, post_data)
    
    print("\n--- COMPARISON RESULT (Color Codes) ---")
    print(json.dumps(comparison, indent=2))
    
    print("\n✓ Phase 2 Complete: Comparison generated successfully")
    
    # Test Phase 3
    print("\n" + "=" * 70)
    print("PHASE 3: Exporting to Excel")
    print("=" * 70)
    
    export_to_excel(pre_data, post_data, comparison, "network_comparison.xlsx")
    
    print("\n" + "=" * 70)
    print("ALL PHASES COMPLETE!")
    print("=" * 70)
    print("\nSummary:")
    for command_name, comp_data in comparison.items():
        overall_color = determine_overall_color(comp_data)
        status = "✓ IDENTICAL" if overall_color == "green" else "✗ CHANGED"
        print(f"  {command_name}: {status} ({overall_color.upper()})")